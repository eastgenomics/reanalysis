#!usr/bin/env python

"""
oc.<client>.search()

-gets data on ALL instances in that client
-e.g. oc.projects.search() gets data on all projects

oc.<client>.info(<client> = <client instance id>)

-gets data on a SUBSET of instances in that client
-e.g. oc.projects.info(projects = <project id>) gets data from one project
"""


import gzip
import json
import pandas as pd
import requests
import subprocess

from bs4 import BeautifulSoup as bs
from datetime import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from panelapp import api, Panelapp, queries

from pyopencga.opencga_config import ClientConfiguration
from pyopencga.opencga_client import OpencgaClient


""" utility functions """


def gzip_file(fp, content):
    with gzip.open(fp, 'wt') as writer:
        writer.write(content)


def ungzip_file(fp):
    with gzip.open(fp, 'rt') as reader:
        data = reader.read()


def generate_date():

    current_date = dt.today()
    date = str(dt.strftime(current_date, "%Y%m%d"))

    return date


def get_ci_panels(td_pa_map, r_code):
    """ Identify which PA panels are currently associated with a CI R
    code using the national genomic test directory.

    args:

    returns:
        panels [list]: list of PA panel IDs or HGNC IDs, or null
    """

    panels = None

    with open(td_pa_map) as reader:
        json_data = json.load(reader)

    for ci in json_data['indications']:
        if ci['code'] == r_code:
            panels = ci['panels']

    return panels


def import_hgnc_dump(csv_file):
    """ Read in a tab-separated dump of the HGNC database and create a
    pandas dataframe from it. The original dump should contain only the
    following 4 columns, in this order:

    'HGNC ID'
    'Approved symbol'
    'Previous symbols'
    'Alias symbols'

    args:
        csv_file [path]: containing dump of HGNC database

    returns:
        hgnc_df [pandas dataframe]
    """

    hgnc_df = pd.read_csv(csv_file, sep='\t')

    hgnc_df.columns = [
        'hgnc_id',
        'symbol',
        'prev_symbol',
        'alias_symbol']

    return hgnc_df


def get_hgnc_from_symbol(hgnc_df, gene_symbol):
    """ Get the HGNC ID for a supplied gene symbol, if one exists. A
    gene symbol may appear in the 'symbol', 'alias_symbol' or
    'prev_symbol' columns depending on whether it is current or not, so
    the function looks through all three in turn.

    args:
        hgnc_df: pandas df containing dump of HGNC site
        gene_symbol [str]: query gene symbol

    returns:
        hgnc_id [str], or None: HGNC ID of query gene
    """

    hgnc_id = None

    # if a row exists where this gene is the official gene symbol,
    # get that row's index and hgnc id

    try:

        target_index = hgnc_df.index[hgnc_df['symbol'] == gene_symbol]
        hgnc_id = hgnc_df.loc[target_index[0], 'hgnc_id']

    except IndexError:

        # or see if it's in the 'previous symbols' field

        try:
            i = 0

            for value in hgnc_df['prev_symbol']:
                if gene_symbol in str(value):

                    hgnc_id = hgnc_df.iloc[i].loc['hgnc_id']
                    break

                i += 1

        # or see if it's in the 'alias symbols' field

        except IndexError:

            j = 0

            for value in hgnc_df['alias_symbol']:
                if gene_symbol in str(value):

                    hgnc_id = hgnc_df.iloc[j].loc['hgnc_id']
                    break

                j += 1

    return hgnc_id


def chromosome_notation(chr_file, vcf_file):
    """ Convert RefSeq contig notation into UCSC format.

    Args:
        chr_file [str]: lookup file for contig/scaffold aliases
        vcf_file [str]: input VCF file

    """

    # Generate a list of contig name conversions

    chr_names = []

    with open(chr_file, 'r') as reader:
        lines = reader.readlines()

    for line in lines:
        elements = line.split('\t')

        values = [
            elements[0].strip(),  # RefSeq contig name
            elements[1].strip(),  # UCSC contig name
            elements[2].strip(),  # Alternative sequence name
            ]

        chr_names.append(values)

    # Use sed to find and replace strings in the target file

    for names in chr_names:

        # Some contigs don't have a UCSC name, so use the alternative one
        if names[1] == 'na':

            subprocess.run([
                'sed',
                '-i',  # change file in-place
                f"s/{names[0]}/{names[2]}/g",
                vcf_file,
                ])

        # Otherwise use the UCSC name
        else:

            subprocess.run([
                'sed',
                '-i',
                f"s/{names[0]}/{names[1]}/g",
                vcf_file,
                ])

    return vcf_file


def format_excel(filename, rows):
    """ Visually format an existing excel file. Apply a style to
    dataframe header cells, and set column widths to autofit data.

    args:
        filename: path to file
    """

    # load in the excel file

    wb = load_workbook(filename=filename)
    ws = wb['Sheet1']

    # define a style for a default font type and size

    normal_font = NamedStyle(name="normal_font")

    normal_font.font = Font(name='Arial', size=10)

    normal_font.alignment = Alignment(
        horizontal='left',
        vertical='center')

    wb.add_named_style(normal_font)

    # define a style to highlight header cells

    highlight = NamedStyle(name="highlight")

    highlight.font = Font(bold=True, name='Arial', size=10)

    highlight.alignment = Alignment(vertical='center')

    highlight.fill = PatternFill(
        fill_type='solid',
        start_color='00C0C0C0',  # light grey
        end_color='00C0C0C0')

    wb.add_named_style(highlight)

    # apply the default formattings

    for col in 'ABCDEFGHIJKLMNOP':
        for row in range(rows + 1):

            ws[f'{col}{row + 1}'].style = 'normal_font'

        ws[f'{col}1'].style = 'highlight'

    # set all columns to be 5cm wide

    for col in 'BGHKLOP':
        ws.column_dimensions[col].width = 25.5

    for col in 'ACDEFIJMN':
        ws.column_dimensions[col].width = 12

    wb.save(filename=filename)


""" high-level opencga functions """


def session_setup(user, host):
    """ Define username and host server to use for OpenCGA login. Log in
    generates password prompt in terminal. Session token then generated and
    used to define OpenCGA client instance for the program.

    args:
        user [string]: username for OpenCGA login
        host [string]: host server URL

    returns:
        oc: OpenCGA client instance based on user credentials and token
    """

    print(f'{dt.now()} Setting up OpenCGA client')

    config_dict = {'rest': {'host': host}}
    config = ClientConfiguration(config_dict)

    oc = OpencgaClient(config)
    oc.login(user)

    token = oc.token
    oc = OpencgaClient(configuration=config, token=token)

    return oc


def get_all_project_ids(oc):
    """ Print and return the ids of all OpenCGA projects to which user
    has access.

    args:
        oc: OpenCGA client

    returns:
        project_ids [list]
    """

    all_projects = oc.projects.search().get_results()

    project_ids = [project['id'] for project in all_projects]
    project_ids.sort()

    for id in project_ids:
        print(id)

    return project_ids


def get_all_study_ids(oc, project_id):
    """ Print and return the ids of all studies in a single project.

    args:
        oc: OpenCGA client
        project_id [str]

    returns:
        study_ids [list]
        """

    project_studies = oc.studies.search(project=project_id).get_results()

    study_ids = [study['id'] for study in project_studies]
    study_ids.sort()

    for id in study_ids:
        print(id)

    return study_ids


def get_all_case_ids(oc, study_id):
    """ Return and write out the ids of all cases in a single study.

    args:
        oc: OpenCGA client
        study_id [str]: id specifying project and study

    returns:
        case_ids [list]: list of case ID strings
    """

    cases = oc.clinical.search(study=study_id, include='id').get_results()

    case_ids = [case['id'] for case in cases]
    case_ids.sort()

    print(f'Study {study_id} contains {len(case_ids)} cases')

    all_case_ids_file = 'do_not_upload/info_lists/all_case_ids.txt'

    with open(all_case_ids_file, 'w') as writer:
        for id in case_ids:

            writer.write(f'{id}\n')

    return case_ids


def get_single_case_data(oc, study_id, case_id):
    """ Retrieve all data from specified case in specified study

    args:
        oc: OpenCGA client
        study_id [str]: id specifying project and study
        case_id [str]: id specifying case within study

    returns:
        case_data [dict]: all data associated with that case
    """

    print(f'{dt.now()} Retrieving case data')

    case_data = oc.clinical.info(
        study=study_id,
        clinical_analysis=case_id
        ).get_results()[0]

    return case_data


def get_multi_case_data(oc, study_id, case_limit):
    """ Retrieve the first <case_limit> cases from a study

    args:
        oc: OpenCGA client
        study_id [str]: id specifying project and study
        case_id [str]: id specifying case within that study

    returns:
        cases: list of case dicts
    """

    cases = oc.clinical.search(
        study=study_id,
        limit=case_limit
        ).get_results()

    return cases


def write_single_case(oc, study_id, case_id):
    """ Retrieve and write out data from a single case.

    args:
        oc: OpenCGA client
        study_id [str]: id specifying project and study
        case_id [str]
    """

    case_file = f'do_not_upload/{case_id}_case_data.json'
    case_data = get_single_case_data(oc, study_id, case_id)

    with open(case_file, 'w') as writer:
        json.dump(case_data, writer)


""" functions on an opencga case """


def get_case_disorder(case):
    """ Identify the main disorder associated with a case.

    args:
        case [dict]

    returns:
        disorder [dict], or None
    """

    if case['disorder']:
        return case['disorder']

    else:
        return None


def get_case_panels(case_data):
    """ Retrieve id, name and version of all panels associated with a case.

    args:
        case_data [dict]: all data from a single case in an OpenCGA study

    returns:
        case_panels [dict]: lists of panel ids, names and versions
    """

    print(f'{dt.now()} Identifying case panels')

    case_panels = {'id': [], 'name': [], 'original_v': []}

    for panel in case_data['panels']:

        case_panels['id'].append(panel['source']['id'])
        case_panels['name'].append(panel['source']['name'])
        case_panels['original_v'].append(panel['source']['version'])

    return case_panels


def print_case_panels(case):
    """ Print the id, name and version of each panel in a case

    args:
        case: dict representing one case in an OpenCGA study
    """

    for panel in case['panels']:

        id = panel['source']['id']
        name = panel['source']['name']
        version = panel['source']['version']

        print(f'{id}: {name} v{version}')


def get_all_case_entities(hgnc_df, case_panels):
    """ Given the panel ids and versions originally used in a case,
    identify all unique green entities across all original or current
    panel versions.

    args:
        hgnc_df [pandas df]: containing dump of hgnc data
        case_panels [dict]: {id, name, original_v} with lists of panel info

    returns:
        case_panels [dict]: updated with green entities
            {id, name, original_v, current_v, original_ents, current_ents}
    """

    print(f'{dt.now()} Identifying all panel entities')

    case_panels['current_v'] = []
    case_panels['original_ents'] = []
    case_panels['current_ents'] = []

    for i in range(len(case_panels['id'])):

        original_panel = get_panelapp_panel(
            case_panels['id'][i], case_panels['original_v'][i])

        current_panel = get_panelapp_panel(case_panels['id'][i])

        case_panels['current_v'].append(current_panel['version'])

        case_panels['original_ents'].append(get_all_green_entities(
            hgnc_df, original_panel))

        case_panels['current_ents'].append(get_all_green_entities(
            hgnc_df, current_panel))

    # check that all dict values are the same length

    for key, value in case_panels.items():

        assert len(value) == len(case_panels['id']), \
            f"case_panels[{key}] is a different length to case_panels['id']"

    return case_panels


def generate_case_summary(date, case_id, case_panels):
    """ Produce a text file detailing the panels used in the case, the
    original entities covered by those panels, and the entities covered
    by the current versions of the same panels.

    args:
        date [str]
        case_id [str]
        case_panels_df [pandas df]: info on panels used in that case
    """

    print(f'{dt.now()} Creating case summary output XLSX')

    filename = f'do_not_upload/{case_id}_entities_{date}.xlsx'

    panel_info = compare_entity_lists(case_panels)

    df = pd.DataFrame(panel_info)

    with pd.ExcelWriter(filename) as writer:
        df.to_excel(writer, index=False)

    format_excel(filename, len(panel_info['id']))


def get_all_case_intpns(case):
    """ Return a list of all interpretation dicts in a case.

    args:
        case [dict]: all data from a single case in an OpenCGA study

    returns:
        all_intpns: list of dicts, each dict is one case interpretation
    """

    all_intpns = []

    all_intpns.append(case['interpretation'])

    for interpretation in case['secondaryInterpretations']:
        all_intpns.append(interpretation)

    return all_intpns


def get_case_phenotypes(case):
    """ Return a list of the 'phenotypes' associated with a case. These
    are identified by iterating over all 'Evidences' in all
    'PrimaryFindings' (variants) in all 'Interpretations' of the case.

    Note that phenotypes are not reliably either HPO terms or names of
    PanelApp panels, because that would be too easy.

    args:
        case [dict]: all data from a single case in an OpenCGA study

    returns:
        phenotypes [list]: each element is one phenotype [str]
    """

    phenotypes = []
    interpretations = get_all_case_intpns(case)

    for intpn in interpretations:
        for variant in intpn['primaryFindings']:
            for evidence in variant['evidences']:
                for phenotype in evidence['phenotypes']:

                    phen_string = phenotype['id']
                    phen_list = phen_string.split(',')

                    for phen in phen_list:
                        if phen.strip() and (phen.strip() not in phenotypes):

                            phenotypes.append(phen.strip())

    phenotypes.sort()

    return phenotypes


def get_variant_dict(variant):
    """ For a single variant, in a single interpretation of an OpenCGA
    case, get a dict of the variant's id, affected genes, and tiers.

    Each variant has as many 'evidences' as its parent interpretation
    has panels. Each evidence is associated with an affected gene and
    the variant's assigned tier. Therefore, a single variant can have
    multiple affected genes and assigned tiers. A variants' tiers -may-
    all be the same, but not necessarily.

    args:
        variant: dict representing a single variant from a single
            interpretation in an OpenCGA case

    returns:
        var_genes: list of unique genes affected by that variant
        var_tiers: list of unique tiers across all variant evidences
    """

    var_genes = []
    var_tiers = []

    for evidence in variant['evidences']:
        if evidence['genomicFeature']['geneName'] not in var_genes:
            var_genes.append(evidence['genomicFeature']['geneName'])

        if evidence['classification']['tier'] not in var_tiers:
            var_tiers.append(evidence['classification']['tier'])

    return var_genes, var_tiers


def compare_entity_lists(case_panels):
    """ Compare the original and current entities for a case, return the
    diff.

    args:
        case_panels []

    returns:
        bool for 'original and current entities are the same'
    """

    # initialise output dict keys

    panel_info = {key: case_panels[key] for key in [
        'id', 'name', 'original_v', 'current_v']}

    panel_info['original_genes'] = []
    panel_info['current_genes'] = []
    panel_info['genes_lost'] = []
    panel_info['genes_gained'] = []
    panel_info['original_regions'] = []
    panel_info['current_regions'] = []
    panel_info['regions_lost'] = []
    panel_info['regions_gained'] = []
    panel_info['original_strs'] = []
    panel_info['current_strs'] = []
    panel_info['strs_lost'] = []
    panel_info['strs_gained'] = []

    # look at the original and current entities for each case panel

    for i in range(len(case_panels['id'])):

        original_ents = case_panels['original_ents'][i]
        current_ents = case_panels['current_ents'][i]

        # look at each entity type and the name of each entity
        # (can't use HGNCs because not all have associated genes)

        for type in 'genes', 'regions', 'strs':

            original = [item[1] for item in original_ents[type]]
            current = [item[1] for item in current_ents[type]]

            panel_info[f'original_{type}'].append(len(original))
            panel_info[f'current_{type}'].append(len(current))

            gained = [item for item in current if item not in original]
            lost = [item for item in original if item not in current]

            if gained:
                gained.sort()
                panel_info[f'{type}_gained'].append(gained)

            else:
                panel_info[f'{type}_gained'].append(None)

            if lost:
                lost.sort()
                panel_info[f'{type}_lost'].append(lost)

            else:
                panel_info[f'{type}_lost'].append(None)

    return panel_info


""" functions on panelapp panels """


def get_panelapp_panel(panel_id, panel_version=None):
    """ Retrieve panel object representing specified version of a
    PanelApp panel ('Panelapp.Panel' doesn't always work, because some
    older panel versions don't contain the hgnc_symbol element)

    args:
        id [str/int]: the panel's PanelApp ID

        version [str/float] (optional): version of the panel to use - if
            not supplied, retrieves current version

    returns:
        panel: dict of data for specified version of specified panel
    """

    # easiest way is to use the panelapp package

    try:
        result = Panelapp.Panel(str(panel_id), panel_version).get_data()

    # except it doesn't work for some older panel versions

    except KeyError:
        path = ["panels", str(panel_id)]
        param = {"version": panel_version}

        url = api.build_url(path, param)
        result = api.get_panelapp_response(url)

    except AttributeError:
        path = ["panels", str(panel_id)]
        param = {"version": panel_version}

        url = api.build_url(path, param)
        result = api.get_panelapp_response(url)

    return result


def write_single_panel_text(id, version=None):
    """ Dump out the contents of a PanelApp panel to a text file

    args:
        id [str/int]: PanelApp ID for a panel
        version [str/float] (optional): specific panel version to get
    """

    panel = get_panelapp_panel(id, version)

    filename = f"panel_dump_{id}_v{panel['version']}.txt"

    with open(filename, 'w') as writer:
        writer.write(str(panel))


def write_single_panel_json(id, version=None):
    """ Dump out the contents of a PanelApp panel to a json file

    args:
        id [str/int]: PanelApp ID for a panel
        version [str/float] (optional): specific panel version to get
    """

    panel = get_panelapp_panel(id, version)

    filename = f"panel_dump_{id}_v{panel['version']}.json"

    with open(filename, 'w') as writer:
        json.dump(panel, writer)


def get_all_green_entities(hgnc_df, panel_data):
    """ Given the data for a single PanelApp panel, return dataframes of
    specific information on each green gene, region and str in that panel.

    args:
        hgnc_df: pandas df containing dump of HGNC site
        panel_data [dict]: for a single PanelApp panel

    returns:
        entity_dfs [dict]: {'genes': list, 'regions': list, 'strs': list}
    """

    panel_entities = {'genes': [], 'regions': [], 'strs': []}

    for key in panel_entities.keys():
        for entity in panel_data[f'{key}']:
            if entity['confidence_level'] == '3':

                info = get_entity_list(hgnc_df, entity)
                panel_entities[key].append(info)

    return panel_entities


def get_entity_list(hgnc_df, entity):
    """ Given an entity (data for a single gene/region/str) from a
    PanelApp panel, extract specific details and return them as a list.

    args:
        hgnc_df: pandas df containing dump of HGNC site
        entity [dict]: single gene/region/STR from a PA panel object

    returns:
        info [list]: details about that entity as [type, name, gene, hgnc]
    """

    # Initialise output

    info = [entity['entity_type']]

    # Get human-readable entity name (different for regions)

    if entity['entity_type'] == 'region':
        info.append(entity['verbose_name'])

    else:
        info.append(entity['entity_name'])

    # get gene symbol and HGNC id (if there is an associated gene)

    if entity['gene_data']:

        gene = entity['gene_data']['gene_symbol']
        hgnc = get_hgnc_from_symbol(hgnc_df, gene)

        info.append(gene)
        info.append(hgnc)

    else:
        info += [None, None]

    return info


def get_hgnc_string(entity_list):
    """ Given data for a single PA panel, construct a string of the
    comma-separated HGNC numbers (not complete HGNC ids) associated with
    those entities (where such exist).

    args:
        panels [list]: PA panel IDs and versions

    returns:
        hgnc_string [str]: comma-separated HGNC numbers
    """

    hgnc_string = ''

    for single_panel in entity_list:
        for entity_type in single_panel.keys():
            for single_entity_list in single_panel[entity_type]:

                if single_entity_list[-1]:  # entity's HGNC id may be None

                    hgnc = single_entity_list[-1][5:]
                    hgnc_string += f'{hgnc},'

    return hgnc_string


def get_ci_panels(td_pa_map, r_code):
    """ Identify which PA panels are currently associated with a CI R
    code using the national genomic test directory.

    args:

    returns:
        panels [list]: list of PA panel IDs or HGNC IDs, or None
    """

    panels = None

    with open(td_pa_map) as reader:
        json_data = json.load(reader)

    for ci in json_data['indications']:
        if ci['code'] == r_code:
            panels = ci['panels']

    return panels


""" functions for creating bed files """


def create_strings(panel_list):
    """ Get all HGNC IDs associated with a list of PA panels """

    hgncs = []
    other_regions = []

    if panel_list:
        for element in panel_list:

            if element.startswith('HGNC:'):
                hgncs.append(element[5:])

            else:
                data = get_panelapp_panel(element)
                genes = get_panel_genes(data)
                regions = get_panel_regions(data)

                hgncs += genes
                other_regions += regions

    hgnc_string = ','.join([ele[5:] for ele in hgncs])
    region_string = ','.join(other_regions)

    return hgnc_string, region_string


def get_panel_genes(panel_data):
    """ Given data from a PA panel, identify all gene HGNC IDs and
    return as a list.

    args:
        panel_data [dict]: data from a PA panel

    returns:
        genes [list]: list of gene HGNC IDs
    """

    genes = []

    for gene in panel_data['genes']:

        if gene['gene_data']['hgnc_id'] and \
            (gene['gene_data']['hgnc_id'] not in genes):

            genes.append(gene['gene_data']['hgnc_id'])

    return genes


def get_panel_regions(panel_data):
    """ Given data from a PA panel, identify all standalone genomic
    regions covered by that panel and return as a list.

    args:
        panel_data [dict]: data from a PA panel

    returns:
        regions [list]: in the form in the form chrom:start:end:strand
    """

    regions = []

    for region in panel_data['regions']:

        coords = region['grch38_coordinates']

        string = f"{region['chromosome']}:{coords[0]}:{coords[1]}"

        regions.append(string)

    return regions


def create_bed(case_id, hgnc_string):
    """ Using the output of get_case_panels(), create bed files for the
    entities (genes and STRs) covered by the case's panels.

    Bed files are created by querying Ensembl BioMart (GRCh38) with
    lists of HGNC IDs.

    args:
        case_id [str]
        case_panels [dict]

    returns:
        fp to bed file
    """

    # given an HGNC ID, how does BioMart turn that into a bed file region?
    # does the region include all exons and introns? what about padding?

    print(f'{dt.now()} Creating bed file')

    # define destination for output bed file

    arg_dict = {
        'input': hgnc_string,
        'output': f'do_not_upload/bed_files/{case_id}.bed'}

    # query BioMart API using subprocess

    with open('bed_text.txt', 'r') as reader:
        contents = reader.read()

    query = contents.replace('PLACEHOLDER', arg_dict['input'])

    biomart_call = [
                    'wget',
                    '-O',
                    f"{arg_dict['output']}",
                    f'{query}',
                    ]

    subprocess.run(biomart_call)

    # sort each file by chromosome and start position

    sort_bed(arg_dict['output'])

    return arg_dict['output']


def sort_bed(filename):
    """ Given the path to a bed file with three headerless tab-delimited
    columns (chrom, start, and end), sort data by 'chrom' then 'start'.

    args:
        filename [str]
    """

    with open(filename, 'r') as reader:
        bed_data = pd.read_csv(reader, sep='\t', header=None)

    bed_data.columns = ['chrom', 'start', 'end']
    sorted = bed_data.sort_values(by=['chrom', 'start'])

    with open(filename, 'w') as writer:
        sorted.to_csv(writer, sep='\t', header=False, index=False)


""" functions on vcf files """


def examine_vcf_header(vcf, output_fp):
    """ Write out just the header from a VCF into a text file you can
    open without it crashing """

    with open(vcf, 'r') as reader:
        lines = reader.readlines()

    output = [line for line in lines if line[0] == '#']

    with open(output_fp, 'w') as writer:
        for line in output:
            writer.write(line)


def liftover_vcf(vcf_file, liftover_vcf, rejected_vcf):
    """ Liftover a GRCh37 VCF file to GRCh38 using Picard LiftoverVcf (GATK).

    args:
        vcf_file
        liftover_vcf
        rejected_vcf
    """

    print(f'{dt.now()} Lifting VCF over to GRCh38')

    subprocess.run([
        './gatk LiftoverVCF',
        f'I={vcf_file}',
        f'O={liftover_vcf}',
        'CHAIN=hg19ToHg38.over.chain',
        f'REJECT={rejected_vcf}',
        'R=GCF_000001405.40_GRCh38.p14_genomic.fna.gz'])


def filter_vcf(vcf_file, bed_file, minQ, output_prefix):
    """  Use vcftools to filter variants in a VCF file on QUAL value and
    a bed file.

    Args:
        vcf_file [str]: input file to be filtered
        bed_file [str]: listing regions to filter on
        minQ [str]: minimum QUAL value to retain a variant
        output_prefix [str]: prefix for output files
    """

    print(f'{dt.now()} Filtering VCF file')

    subprocess.run([
        'vcftools',
        '--vcf', vcf_file,
        '--bed', bed_file,
        '--minQ', minQ,
        '--recode',
        '--out', output_prefix,
        ])

    return f'{output_prefix}.recode.vcf'


def sort_vcf(vcf_file, output_fp):
    """ Use bcftools to sort a VCF file.

    Args:
        vcf_file [str]: input VCF file
        output_fp [str]: path to sorted output file
    """

    print(f'{dt.now()} Sorting VCF file')

    # ./gatk SortVcf /
    # I=initial_vcf /
    # O=sorted_vcf

    subprocess.run([
        'bcftools',
        'sort',
        '-o', output_fp,
        vcf_file,
        ])

    return output_fp


def annotate_vcf(vcf_file, output_fp):
    """ Use the CellBase API to annotate a VCF file.

    Args:
        vcf_file [str]: input VCF file
        output_fp [str]: path to sorted output file
    """

    print(f'{dt.now()} Annotating VCF file')

    variants = get_variant_list(vcf_file)

    data = get_cellbase_response(variants)


def get_variant_list(vcf_file):
    """ Get a list of all variants in a VCF file.

    args:
        vcf_file: based on GRCh38

    returns:
        variants [str]: in the form 'chrom:pos:ref:alt,chrom:pos:ref:alt,...'
    """


def get_cellbase_response(variants):
    """ Query the CellBase API

    args:
        variants [str]: in the form 'chrom:pos:ref:alt,chrom:pos:ref:alt,...'

    returns:
        data
    """

    suffix = f'/v5/hsapiens/genomic/variant/{variants}/annotation'
    param = 'assembly=grch38'
    query = "{}?{}".format(suffix, param)

    url = f'https://ws.zettagenomics.com/cellbase/webservices/rest/{query}'

    for i in range(0, 5):

        try:
            request = requests.get(url, headers={"Accept": "application/json"})

        except Exception as error:

            print("Error retrieving data from CellBase: {}".format(error))
            return None

        else:
            if request.ok:

                data = json.loads(request.content.decode("utf-8"))
                return data

            else:

                print("Error {} for URL: {}".format(request.status_code, url))
                return None


def enumerate_variants(fp):
    """ Given the path to a VCF file, count:

    - The total number of lines
    - The number of variant (non-header) lines
    - The number of SNV (single-base REF and ALT) lines
    """

    with open(fp, 'r') as reader:
        lines = reader.readlines()

    total_lines = 0
    non_header_lines = 0
    snv_lines = 0

    for line in lines:

        total_lines += 1

        if line[0] != '#':

            non_header_lines += 1

            values = [value.strip() for value in line.split('\t')]

            if (len(values[3]) == 1) and (len(values[4]) == 1):

                snv_lines += 1

    print(f'{fp}\n'
        f'Total lines: {total_lines}\n'
        f'Variants: {non_header_lines}\n'
        f'SNVs: {snv_lines}\n')


""" looking at data across an opencga study """


def check_single_case_panels(case):
    """ Check whether 'case > panels' contains the same panels as are
    covered by all the case's interpretations

    args:
        case: dict representing one case in an OpenCGA study

    returns:
        panels_match [bool]: True/False for whether "all panels in
            case>panels" and "all panels contained within the case's
            interpretations" are equivalent
    """

    # Get list of panels in 'case > panels'

    case_panels = []

    for panel in case['panels']:

        case_panels.append({
            'id': panel['source']['id'],
            'name': panel['name'],
            'version': panel['source']['version']})

    # Get list of all panels within all case interpretations

    interpretations = get_all_case_intpns(case)

    all_intpn_panels = []

    for intpn in interpretations:
        for panel in intpn['panels']:

            all_intpn_panels.append({
                'id': panel['source']['id'],
                'name': panel['name'],
                'version': panel['source']['version']})

    # Check whether each intpn panel is in the case panels list

    intpn_panel_count = len(all_intpn_panels)
    matches = 0

    for panel in all_intpn_panels:
        if panel in case_panels:
            matches += 1

    if matches == intpn_panel_count:
        return True

    else:
        return False


def check_panels_all_cases(oc, study_id):
    """ Iterate over every case in a study and check that the panels
    covered in 'case > panels' are the same as those covered by
    'interpretation > panels' across all interpretations

    args:
        oc: OpenCGA client
        study_id [str]: id specifying project and study

    returns:
        all_panels_match [bool]: True/False for whether all cases
            satisfy "'case > panels' and 'interpretations > panels' are
            equivalent"
    """

    case_ids = get_all_case_ids(oc, study_id)

    case_count = len(case_ids)
    matching_cases = 0

    for case_id in case_ids:

        case = get_single_case_data(oc, study_id, case_id)
        check = check_single_case_panels(case)

        if check:
            matching_cases += 1

    if matching_cases == case_count:
        print('Case panels match interpretation panels for all cases')

    else:
        mismatches = case_count - matching_cases
        print(f'Panel mismatches in {mismatches} of {case_count} cases')


def print_awkward_variant(oc):
    """ Print out the data for the awkward variant which has two
    different tiers """

    case_info = oc.clinical.info(
        study='emee-glh@reanalysis:rd37',
        clinical_analysis='SAP-3904-1'
        ).get_results()[0]

    variants = case_info['interpretation']['primaryFindings']

    awkward_variant = '2:47394827:CGTCTCA:C'

    for variant in variants:
        if variant['id'] == awkward_variant:
            for key, value in variant.items():
                print(f'{key}: {value}')


""" Implementation """


def main():

    """ Case details and session setup """

    # date = generate_date()

    # user = 'jmiles'
    # host = 'https://uat.eglh.app.zettagenomics.com/opencga'

    # project = 'reanalysis'
    # study = 'rd37'
    # study_id = f'emee-glh@{project}:{study}'

    # # set up an OpenCGA session
    # oc = session_setup(user, host)

    """ Reference data """

    all_case_ids_file = 'all_case_ids.txt'

    folder = 'do_not_upload/vcfs/'
    case_id = 'SAP-48034-1'
    case_prefix = f'{folder}{case_id}'

    initial_vcf = f'{folder}{case_id}.vcf'
    liftover_vcf = f'{case_prefix}_1_lifted.vcf'
    rejected_vcf = f'{case_prefix}_1_lifted.vcf.unmap'
    filtered_vcf = f'{case_prefix}_2_filtered.recode.vcf'
    sorted_vcf = f'{case_prefix}_3_sorted.vcf'
    annotated_vcf = f'{case_prefix}_annotated.vcf'
    bed_file = f'do_not_upload/bed_files/{case_id}.bed'

    hgnc_dump = '20220817_hgnc_dump.txt'
    # hgnc_df = import_hgnc_dump(hgnc_dump)

    """ Store OpenCGA data locally """

    # # retrieve and write out ids of all cases
    # all_case_ids = get_all_case_ids(oc, study_id)

    # # retrieve and write out data from all cases (THIS IS 1.8 GB BEWARE)
    # write_all_study_cases(oc, study_id, all_case_ids)

    # # retrieve and write out data from specific case (SAP-48034-1)
    # write_single_case(oc, study_id, case_id)

    """ Read in stored data """

    # # read in all case ids
    # with open(all_case_ids_file, 'r') as reader:
    #     all_case_ids = [line.strip() for line in reader.readlines()]

    # # read in all case data
    # with open(all_cases_file, 'r') as reader:
    #     all_case_data = json.load(reader)

    # # read in single case data (SAP-48034-1)
    # with open(single_case_file) as reader:
    #     case_data = json.load(reader)

    """ whatever I'm currently testing """

    fp = 'do_not_upload/vcfs/SAP-48034-1_1_proband.vcf.gz'

    with open('do_not_upload/vcfs/SAP-48034-1_1_proband.vcf', 'r') as reader:
        content = reader.read()

    gzip_file(fp, content)



if __name__ == '__main__':
    main()
